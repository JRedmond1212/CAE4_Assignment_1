import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
from io import BytesIO
import plotly.graph_objects as go
import os
import hashlib
import openpyxl
from typing import Tuple, List, Optional
from math import comb

# =========================================================
# Files / constants
# =========================================================

# This should be the Excel you attached (used as the OUTPUT format template)
DESIGN_TABLE_TEMPLATE_PATH = "/mnt/data/blade_hub_cfd_design_table (1).xlsx"
DESIGN_TABLE_SHEET_NAME = "DesignTable"

# Config template filename (download)
CONFIG_TEMPLATE_FILENAME = "config_template.xlsx"

COLOR_ROOT_C1 = "#1f77b4"   # blue
COLOR_ROOT_C2 = "#d62728"   # red
COLOR_TIP_C1  = "#2ca02c"   # green
COLOR_TIP_C2  = "#ff7f0e"   # orange
COLOR_HUB     = "#9467bd"   # purple

# =========================================================
# UX: keep scroll position on rerun
# =========================================================

def keep_scroll_position():
    components.html(
        """
        <script>
        (function() {
          const key = "st_scrollY";
          function saveScroll() {
            try { sessionStorage.setItem(key, String(window.scrollY || 0)); } catch(e) {}
          }
          function restoreScroll() {
            try {
              const y = parseInt(sessionStorage.getItem(key) || "0", 10);
              if (!isNaN(y)) window.scrollTo(0, y);
            } catch(e) {}
          }
          window.addEventListener("beforeunload", saveScroll);
          window.addEventListener("scroll", function(){ saveScroll(); }, {passive:true});
          setTimeout(restoreScroll, 50);
        })();
        </script>
        """,
        height=0,
    )

# =========================================================
# Maths helpers
# =========================================================

def binomial_coeff(n, k):
    return comb(n, k)

def general_bezier_curve(control_points, num_points=201):
    cp = np.asarray(control_points, dtype=float)
    if cp.ndim != 2 or cp.shape[1] != 2:
        raise ValueError("control_points must be (N,2) array.")
    n = cp.shape[0]
    degree = n - 1
    if degree < 1:
        raise ValueError("Need >=2 control points.")
    u = np.linspace(0.0, 1.0, num_points)
    B = np.zeros((n, len(u)))
    for i in range(n):
        coeff = binomial_coeff(degree, i)
        B[i, :] = coeff * (u ** i) * ((1 - u) ** (degree - i))
    a = np.dot(cp[:, 0], B)
    b = np.dot(cp[:, 1], B)
    return u, a, b

def bezier_point_and_tangent(control_points, u):
    cp = np.asarray(control_points, dtype=float)
    if cp.ndim != 2 or cp.shape[1] != 2:
        raise ValueError("control_points must be (N,2) array.")
    n = cp.shape[0]
    degree = n - 1
    if degree < 1:
        raise ValueError("Need >=2 control points.")
    B = np.zeros(n)
    for i in range(n):
        coeff = binomial_coeff(degree, i)
        B[i] = coeff * (u ** i) * ((1 - u) ** (degree - i))
    point = np.dot(cp.T, B)
    cp_diff = degree * (cp[1:, :] - cp[:-1, :])
    Bp = np.zeros(degree)
    for j in range(degree):
        coeff = binomial_coeff(degree - 1, j)
        Bp[j] = coeff * (u ** j) * ((1 - u) ** (degree - 1 - j))
    tangent = np.dot(cp_diff.T, Bp)
    return point, tangent

def compute_curve2_cubic_c2(root_c1_cubic, trailing_edge=(0.0, 0.0)):
    P1 = root_c1_cubic[1, :]
    P2 = root_c1_cubic[2, :]
    P3 = root_c1_cubic[3, :]
    Q0 = P3
    Q1 = 2.0 * P3 - P2
    Q2 = 4.0 * P3 - 4.0 * P2 + P1
    Q3 = np.array(trailing_edge, dtype=float)
    return np.vstack([Q0, Q1, Q2, Q3])

def transform_tip_from_root(cp_root, scale_x, scale_z, twist_deg_per_m, blade_length, twist_sign=+1.0):
    cp_root = np.asarray(cp_root, dtype=float)
    x_le = np.max(cp_root[:, 0])
    z_le = 0.0
    x_rel = cp_root[:, 0] - x_le
    z_rel = cp_root[:, 1] - z_le

    x_scaled = scale_x * x_rel
    z_scaled = scale_z * z_rel

    twist_total_deg = twist_sign * twist_deg_per_m * blade_length
    theta = np.deg2rad(twist_total_deg)
    cos_t = np.cos(theta)
    sin_t = np.sin(theta)

    x_rot = cos_t * x_scaled + sin_t * z_scaled
    z_rot = -sin_t * x_scaled + cos_t * z_scaled

    x_tip = x_rot + x_le
    z_tip = z_rot + z_le
    y_tip = np.ones_like(x_tip) * blade_length

    cp_tip_2d = np.column_stack([x_tip, z_tip])
    cp_tip_3d = np.column_stack([x_tip, y_tip, z_tip])
    return cp_tip_2d, cp_tip_3d

# =========================================================
# Variant table logic
# =========================================================

def default_variant_table() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "NumBlades": [3, 4, 5],
            "BladeLength_m": [50.0, 40.0, 30.0],
            "Thickness_m": [0.018, 0.015, 0.010],
        }
    )

def parse_variant_table_from_excel(xls: pd.ExcelFile) -> Optional[pd.DataFrame]:
    candidate_sheets = []
    for s in xls.sheet_names:
        name = s.lower()
        if "table1" == name or ("variant" in name) or ("blade" in name and "table" in name):
            candidate_sheets.append(s)
    candidate_sheets = candidate_sheets + [s for s in xls.sheet_names if s not in candidate_sheets]

    def normalise_cols(df: pd.DataFrame) -> pd.DataFrame:
        col_map = {}
        for c in df.columns:
            cl = str(c).strip().lower()
            if cl in ["numblades", "num_blades", "blades", "nblades", "n_blades"]:
                col_map[c] = "NumBlades"
            if cl in ["blade_length", "bladelength", "length", "blade_length_m", "bladelength_m", "length_m"]:
                col_map[c] = "BladeLength_m"
            if cl in ["thickness", "blade_thickness", "bladethickness", "thickness_m", "blade_thickness_m"]:
                col_map[c] = "Thickness_m"
        return df.rename(columns=col_map).copy()

    for s in candidate_sheets:
        try:
            df = pd.read_excel(xls, sheet_name=s)
        except Exception:
            continue
        if df is None or df.empty:
            continue
        df2 = normalise_cols(df)
        if {"NumBlades", "BladeLength_m", "Thickness_m"}.issubset(df2.columns):
            out = df2[["NumBlades", "BladeLength_m", "Thickness_m"]].dropna()
            try:
                out["NumBlades"] = out["NumBlades"].astype(int)
                out["BladeLength_m"] = out["BladeLength_m"].astype(float)
                out["Thickness_m"] = out["Thickness_m"].astype(float)
            except Exception:
                continue
            out = out.drop_duplicates(subset=["NumBlades"]).sort_values("NumBlades")
            if len(out) >= 1:
                return out.reset_index(drop=True)
    return None

def blade_length_and_thickness_from_table(num_blades: int, table: pd.DataFrame) -> Tuple[float, float]:
    sub = table[table["NumBlades"] == int(num_blades)]
    if sub.empty:
        r0 = table.iloc[0]
        return float(r0["BladeLength_m"]), float(r0["Thickness_m"])
    r = sub.iloc[0]
    return float(r["BladeLength_m"]), float(r["Thickness_m"])

# =========================================================
# Excel / config helpers
# =========================================================

def file_hash_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()

def sort_curve_points(df_curve):
    df_curve = df_curve.copy()
    if "PointIndex" in df_curve.columns:
        df_curve = df_curve.sort_values("PointIndex")
    elif "Point" in df_curve.columns:
        def extract_index(p):
            try:
                import re
                m = re.search(r"(\d+)$", str(p))
                if m:
                    return int(m.group(1))
            except Exception:
                pass
            return 999999
        df_curve["__idx__"] = df_curve["Point"].apply(extract_index)
        df_curve = df_curve.sort_values("__idx__").drop(columns="__idx__")
    return df_curve

def load_bezier_control_points(excel_file_like):
    try:
        xls = pd.ExcelFile(excel_file_like)
    except Exception:
        return None
    if "BezierControlPoints" not in xls.sheet_names:
        return None
    return pd.read_excel(xls, sheet_name="BezierControlPoints")

def load_parameters(excel_file_like):
    try:
        xls = pd.ExcelFile(excel_file_like)
    except Exception:
        return {}
    if "Parameters" not in xls.sheet_names:
        return {}
    df = pd.read_excel(xls, sheet_name="Parameters")
    if "Name" not in df.columns or "Value" not in df.columns:
        return {}
    params = {}
    for _, row in df.iterrows():
        name = row["Name"]
        value = row["Value"]
        if pd.isna(name) or pd.isna(value):
            continue
        params[str(name)] = value
    return params

def load_hub_control_points(excel_file_like):
    try:
        xls = pd.ExcelFile(excel_file_like)
    except Exception:
        return None, None
    if "HubControlPoints" not in xls.sheet_names:
        return None, None
    df = pd.read_excel(xls, sheet_name="HubControlPoints")

    used = None
    if {"X", "Z"}.issubset(df.columns):
        used = "X–Z"
        df_sorted = sort_curve_points(df) if ("PointIndex" in df.columns or "Point" in df.columns) else df
        A = df_sorted["X"].to_numpy()
        B = df_sorted["Z"].to_numpy()
    elif {"Y", "Z"}.issubset(df.columns):
        used = "Y–Z"
        df_sorted = sort_curve_points(df) if ("PointIndex" in df.columns or "Point" in df.columns) else df
        A = df_sorted["Y"].to_numpy()
        B = df_sorted["Z"].to_numpy()
    else:
        return None, None

    if np.isnan(A).any() or np.isnan(B).any():
        return None, None
    cp = np.column_stack([A, B])
    if cp.shape[0] < 2:
        return None, None
    return cp, used

# =========================================================
# Streamlit state defaults
# =========================================================

def ensure_state_defaults():
    if "defaults_initialised" in st.session_state:
        return
    st.session_state.defaults_initialised = True

    st.session_state.blade_options = [3, 4, 5]
    st.session_state.num_blades = 3
    st.session_state.scale_x = 0.8
    st.session_state.scale_z = 0.8
    st.session_state.twist_total_deg = 15.0
    st.session_state.twist_sign = +1.0

    st.session_state.c1_p0_x = 1.0
    st.session_state.c1_p0_z = 0.0
    st.session_state.c1_p1_x = 1.0
    st.session_state.c1_p1_z = 0.0600
    st.session_state.c1_p2_x = 0.8
    st.session_state.c1_p2_z = 0.0800
    st.session_state.c1_p3_x = 0.6
    st.session_state.c1_p3_z = 0.0750

    st.session_state.hub_p0_x = -0.4
    st.session_state.hub_p0_z = 1.0
    st.session_state.hub_p1_x = 0.3
    st.session_state.hub_p1_z = 1.0
    st.session_state.hub_p2_x = 1.3
    st.session_state.hub_p2_z = 0.8
    st.session_state.hub_p3_x = 2.0
    st.session_state.hub_p3_z = 0.0

    # Abaqus defaults (UI units)
    st.session_state.abaqus_load_mag_kN = 1.0
    st.session_state.abaqus_u = 0.50
    st.session_state.abaqus_part_name = "Blade"
    st.session_state.abaqus_job_name = "BladeJob"
    st.session_state.abaqus_script_filename = "Blade_abaqus_code.py"

    st.session_state.abaqus_material_name = "Fibreglass"
    st.session_state.abaqus_E_MPa = 70000.0
    st.session_state.abaqus_nu = 0.33
    st.session_state.abaqus_mesh_seed_model_units = 250.0

    st.session_state.last_config_hash = None

    # Config results kept in state
    st.session_state.config_applied = False
    st.session_state.config_warnings = []
    st.session_state.variant_table = default_variant_table()

def apply_config_once_if_new(uploaded_bytes: bytes) -> Tuple[bool, List[str], pd.DataFrame]:
    """
    Must be called from file_uploader callback to avoid mutating widget-backed keys
    after the widgets already exist.
    """
    warnings = []
    applied = False
    variant_table = default_variant_table()

    if uploaded_bytes is None:
        return applied, warnings, variant_table

    cfg_hash = file_hash_bytes(uploaded_bytes)
    if st.session_state.last_config_hash == cfg_hash:
        return applied, warnings, variant_table

    st.session_state.last_config_hash = cfg_hash
    applied = True

    try:
        xls = pd.ExcelFile(BytesIO(uploaded_bytes))
    except Exception:
        warnings.append("Config upload could not be read as an Excel file. Using current sidebar values.")
        return applied, warnings, variant_table

    vt = parse_variant_table_from_excel(xls)
    if vt is not None:
        variant_table = vt
        new_opts = list(map(int, variant_table["NumBlades"].tolist()))
        if len(new_opts) >= 1:
            st.session_state.blade_options = new_opts
            if int(st.session_state.num_blades) not in new_opts:
                st.session_state.num_blades = int(new_opts[0])
    else:
        warnings.append("No blade variant table found in config; using default blade options (3,4,5).")

    params = load_parameters(BytesIO(uploaded_bytes))
    if params:
        def try_set_float(key_state: str, param_name: str):
            if param_name in params:
                try:
                    st.session_state[key_state] = float(params[param_name])
                except Exception:
                    warnings.append(f"Parameter '{param_name}' was not a valid number; ignored.")

        def try_set_int(key_state: str, param_name: str):
            if param_name in params:
                try:
                    st.session_state[key_state] = int(float(params[param_name]))
                except Exception:
                    warnings.append(f"Parameter '{param_name}' was not a valid integer; ignored.")

        try_set_int("num_blades", "NumBlades")
        if st.session_state.blade_options and int(st.session_state.num_blades) not in st.session_state.blade_options:
            st.session_state.num_blades = int(st.session_state.blade_options[0])

        try_set_float("scale_x", "ScaleX")
        try_set_float("scale_z", "ScaleZ")
        try_set_float("twist_total_deg", "TwistTotalDeg")
        if "TwistSign" in params:
            try:
                st.session_state.twist_sign = +1.0 if float(params["TwistSign"]) >= 0 else -1.0
            except Exception:
                warnings.append("Parameter 'TwistSign' invalid; ignored.")

        # Abaqus: IMPORTANT: includes AbaqusU -> sets abaqus_u
        if "AbaqusLoadMag_kN" in params:
            try_set_float("abaqus_load_mag_kN", "AbaqusLoadMag_kN")
        elif "AbaqusLoadMag" in params:
            try:
                st.session_state.abaqus_load_mag_kN = float(params["AbaqusLoadMag"]) / 1000.0
            except Exception:
                warnings.append("Parameter 'AbaqusLoadMag' invalid; ignored.")

        if "AbaqusU" in params:
            try_set_float("abaqus_u", "AbaqusU")

        if "AbaqusMaterialName" in params:
            st.session_state.abaqus_material_name = str(params["AbaqusMaterialName"])

        if "AbaqusE_MPa" in params:
            try_set_float("abaqus_E_MPa", "AbaqusE_MPa")
        elif "AbaqusE" in params:
            try:
                st.session_state.abaqus_E_MPa = float(params["AbaqusE"]) / 1e6
            except Exception:
                warnings.append("Parameter 'AbaqusE' invalid; ignored.")

        if "AbaqusNu" in params:
            try_set_float("abaqus_nu", "AbaqusNu")

        if "AbaqusMeshSeed_ModelUnits" in params:
            try_set_float("abaqus_mesh_seed_model_units", "AbaqusMeshSeed_ModelUnits")
        elif "AbaqusMeshSeed" in params:
            try_set_float("abaqus_mesh_seed_model_units", "AbaqusMeshSeed")

        if "AbaqusPartName" in params:
            st.session_state.abaqus_part_name = str(params["AbaqusPartName"])
        if "AbaqusJobName" in params:
            st.session_state.abaqus_job_name = str(params["AbaqusJobName"])
        if "AbaqusScriptFilename" in params:
            st.session_state.abaqus_script_filename = str(params["AbaqusScriptFilename"])

    cp_df = load_bezier_control_points(BytesIO(uploaded_bytes))
    if cp_df is not None:
        sub = cp_df[(cp_df.get("Section") == "Root") & (cp_df.get("Curve") == 1)]
        if not sub.empty and {"X", "Z"}.issubset(sub.columns):
            sub_sorted = sort_curve_points(sub)
            X = sub_sorted["X"].to_numpy()
            Z = sub_sorted["Z"].to_numpy()
            if not (np.isnan(X).any() or np.isnan(Z).any()) and X.shape[0] >= 4:
                st.session_state.c1_p0_x = float(X[0]); st.session_state.c1_p0_z = float(Z[0])
                st.session_state.c1_p1_x = float(X[1]); st.session_state.c1_p1_z = float(Z[1])
                st.session_state.c1_p2_x = float(X[2]); st.session_state.c1_p2_z = float(Z[2])
                st.session_state.c1_p3_x = float(X[3]); st.session_state.c1_p3_z = float(Z[3])
            else:
                warnings.append("BezierControlPoints Root/Curve1 found but invalid; ignored.")

    hub_cp, _hub_plane = load_hub_control_points(BytesIO(uploaded_bytes))
    if hub_cp is not None and hub_cp.shape[0] >= 4:
        st.session_state.hub_p0_x = float(hub_cp[0, 0]); st.session_state.hub_p0_z = float(hub_cp[0, 1])
        st.session_state.hub_p1_x = float(hub_cp[1, 0]); st.session_state.hub_p1_z = float(hub_cp[1, 1])
        st.session_state.hub_p2_x = float(hub_cp[2, 0]); st.session_state.hub_p2_z = float(hub_cp[2, 1])
        st.session_state.hub_p3_x = float(hub_cp[3, 0]); st.session_state.hub_p3_z = float(hub_cp[3, 1])

    return applied, warnings, variant_table

# =========================================================
# Plotting
# =========================================================

def apply_common_layout_tweaks(fig, title, x_range=None, z_range=None, x_label="x", z_label="z"):
    fig.update_layout(
        title=title,
        xaxis_title=x_label,
        yaxis_title=z_label,
        width=None,
        height=600,
        margin=dict(l=40, r=20, t=60, b=60),
        legend=dict(orientation="h", yanchor="top", y=-0.2, xanchor="center", x=0.5),
    )
    fig.update_yaxes(scaleanchor="x", scaleratio=1)
    fig.update_yaxes(zeroline=True, zerolinewidth=2)
    fig.add_hline(y=0, line_width=1, line_dash="dash", opacity=0.5)
    if x_range is not None:
        fig.update_xaxes(range=x_range)
    if z_range is not None:
        fig.update_yaxes(range=z_range)
    return fig

def compute_axis_ranges(root_c1, root_c2, tip_c1_top_2d, tip_c2_top_2d, tip_c1_bottom_2d, tip_c2_bottom_2d):
    x_list = [root_c1[:, 0], root_c2[:, 0]]
    z_list = [root_c1[:, 1], root_c2[:, 1]]
    for arr in [tip_c1_top_2d, tip_c2_top_2d, tip_c1_bottom_2d, tip_c2_bottom_2d]:
        if arr is not None:
            x_list.append(arr[:, 0])
            z_list.append(arr[:, 1])
    x_all = np.concatenate(x_list)
    z_all = np.concatenate(z_list)
    x_min, x_max = float(np.min(x_all)), float(np.max(x_all))
    z_min, z_max = float(np.min(z_all)), float(np.max(z_all))
    dx = (x_max - x_min) or 1.0
    dz = (z_max - z_min) or 1.0
    margin_x = 0.05 * dx
    margin_z = 0.05 * dz
    return [x_min - margin_x, x_max + margin_x], [z_min - margin_z, z_max + margin_z]

def make_root_and_tip_plot_only(
    root_c1, root_c2,
    tip_c1_top_2d, tip_c2_top_2d,
    tip_c1_bottom_2d, tip_c2_bottom_2d,
    x_range, z_range,
    arrow_head_xz: Tuple[float, float],
    arrow_dir_xz_unit: Tuple[float, float],
    arrow_len: float,
):
    fig = go.Figure()

    u1, x1, z1_top = general_bezier_curve(root_c1)
    u2, x2, z2_top = general_bezier_curve(root_c2)
    fig.add_trace(go.Scatter(x=x1, y=z1_top, mode="lines", name="Root C1 top", line=dict(color=COLOR_ROOT_C1)))
    fig.add_trace(go.Scatter(x=x2, y=z2_top, mode="lines", name="Root C2 top", line=dict(color=COLOR_ROOT_C2)))
    fig.add_trace(go.Scatter(x=x1, y=-z1_top, mode="lines", name="Root C1 bottom", line=dict(color=COLOR_ROOT_C1)))
    fig.add_trace(go.Scatter(x=x2, y=-z2_top, mode="lines", name="Root C2 bottom", line=dict(color=COLOR_ROOT_C2)))

    u1t, x1t, z1t = general_bezier_curve(tip_c1_top_2d)
    u2t, x2t, z2t = general_bezier_curve(tip_c2_top_2d)
    u1b, x1b, z1b = general_bezier_curve(tip_c1_bottom_2d)
    u2b, x2b, z2b = general_bezier_curve(tip_c2_bottom_2d)

    fig.add_trace(go.Scatter(x=x1t, y=z1t, mode="lines", name="Tip C1 top",
                             line=dict(dash="dash", color=COLOR_TIP_C1)))
    fig.add_trace(go.Scatter(x=x2t, y=z2t, mode="lines", name="Tip C2 top",
                             line=dict(dash="dash", color=COLOR_TIP_C2)))
    fig.add_trace(go.Scatter(x=x1b, y=z1b, mode="lines", name="Tip C1 bottom",
                             line=dict(dash="dash", color=COLOR_TIP_C1)))
    fig.add_trace(go.Scatter(x=x2b, y=z2b, mode="lines", name="Tip C2 bottom",
                             line=dict(dash="dash", color=COLOR_TIP_C2)))

    hx, hz = float(arrow_head_xz[0]), float(arrow_head_xz[1])
    dx, dz = float(arrow_dir_xz_unit[0]), float(arrow_dir_xz_unit[1])
    tx = hx - arrow_len * dx
    tz = hz - arrow_len * dz

    fig.update_layout(
        annotations=[
            dict(
                x=hx, y=hz,
                ax=tx, ay=tz,
                xref="x", yref="y",
                axref="x", ayref="y",
                showarrow=True,
                arrowhead=3,
                arrowsize=1.0,
                arrowwidth=2,
                text="",
            )
        ]
    )

    return apply_common_layout_tweaks(
        fig,
        "Blade Outer Profile (Root & Tip, mirrored)",
        x_range=x_range,
        z_range=z_range,
        x_label="x (chord-direction)",
        z_label="z (thickness-direction)",
    )

def make_hub_plot(hub_cp_xz, hub_plane_label="X–Z"):
    u, a_vals, z_vals = general_bezier_curve(hub_cp_xz)
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=a_vals, y=z_vals, mode="lines",
        name="Hub profile",
        line=dict(color=COLOR_HUB),
        customdata=np.column_stack([u]),
        hovertemplate=(
            f"Hub profile ({hub_plane_label})<br>u=%{{customdata[0]:.3f}}<br>"
            f"{hub_plane_label.split('–')[0]}=%{{x:.4f}}<br>Z=%{{y:.4f}}<extra></extra>"
        ),
    ))
    fig.add_trace(go.Scatter(
        x=hub_cp_xz[:, 0], y=hub_cp_xz[:, 1],
        mode="lines+markers",
        name="Hub CPs",
        line=dict(dash="dot", color=COLOR_HUB),
        marker=dict(color=COLOR_HUB),
        hovertemplate=(
            f"Hub CP<br>{hub_plane_label.split('–')[0]}=%{{x:.4f}}<br>Z=%{{y:.4f}}<extra></extra>"
        ),
    ))
    xlab = f"{hub_plane_label.split('–')[0]} (hub plane)"
    return apply_common_layout_tweaks(fig, "Hub external profile", x_label=xlab, z_label="Z")

# =========================================================
# Abaqus script generator
#   IMPORTANT: LOAD_U and LOAD_POINT are embedded from the CURRENT Streamlit slider value.
# =========================================================

def generate_abaqus_post_import_script_refined(
    part_name: str,
    job_name: str,
    script_model_name: str,
    instance_name: str,
    material_name: str,
    E_MPa: float,
    nu: float,
    shell_thickness_m: float,
    mesh_seed_model_units: float,
    load_u: float,
    load_point_m: Tuple[float, float, float],
    force_vec_N: Tuple[float, float, float],
):
    MODEL_NAME = script_model_name
    PART_NAME = part_name
    INSTANCE_NAME = instance_name
    MATERIAL_NAME = material_name
    STEP_NAME = "Step-1"
    E_Pa = float(E_MPa) * 1e6

    # NOTE: This script will partition at LOAD_U and then apply the force at the closest
    # node to LOAD_POINT (which is computed from the SAME u in the Streamlit app).
    return f"""#Code: V1.5.4

from abaqus import mdb
from abaqusConstants import *
import regionToolset
import mesh
import math

MODEL_NAME    = r"{MODEL_NAME}"
PART_NAME     = r"{PART_NAME}"
INSTANCE_NAME = r"{INSTANCE_NAME}"

MATERIAL_NAME = r"{MATERIAL_NAME}"
STEP_NAME     = r"{STEP_NAME}"
JOB_NAME      = r"{job_name}"

E       = {E_Pa:.16g}
NU      = {float(nu):.16g}
SHELL_T = {float(shell_thickness_m):.16g}

MESH_SEED = {float(mesh_seed_model_units):.16g}

# --- USER-DEFINED LOCATION ALONG THE TIP EDGE (0..1)
LOAD_U = {float(load_u):.16g}

# --- LOAD POINT GENERATED FROM THE SAME u VALUE IN THE STREAMLIT APP
LOAD_POINT = ({float(load_point_m[0]):.16g}, {float(load_point_m[1]):.16g}, {float(load_point_m[2]):.16g})
FORCE_VEC  = ({float(force_vec_N[0]):.16g}, {float(force_vec_N[1]):.16g}, {float(force_vec_N[2]):.16g})

AUTO_UNIT_SCALE = True
UNIT_SCALE = 1.0

def log(msg):
    print(">>> " + str(msg))

def die(msg):
    raise RuntimeError(str(msg))

def get_model():
    if MODEL_NAME not in mdb.models:
        die("Model '%s' not found. Available models: %s" % (MODEL_NAME, list(mdb.models.keys())))
    return mdb.models[MODEL_NAME]

def get_part(model):
    if PART_NAME not in model.parts:
        die("Part '%s' not found. Available parts: %s" % (PART_NAME, list(model.parts.keys())))
    return model.parts[PART_NAME]

def get_or_create_instance(model, part_obj):
    a = model.rootAssembly
    if INSTANCE_NAME in a.instances:
        log("Using existing instance: %s" % INSTANCE_NAME)
        return a.instances[INSTANCE_NAME]

    if len(a.instances) > 0:
        inst0 = a.instances.values()[0]
        log("WARNING: Instance '%s' not found; reusing first existing instance: %s" % (INSTANCE_NAME, inst0.name))
        return inst0

    log("No instances found. Creating instance %s." % INSTANCE_NAME)
    a.DatumCsysByDefault(CARTESIAN)
    inst = a.Instance(name=INSTANCE_NAME, part=part_obj, dependent=ON)
    return inst

def bbox_from_part_geometry(part_obj):
    pts = []
    try:
        for v in part_obj.vertices:
            p = v.pointOn[0]
            pts.append((float(p[0]), float(p[1]), float(p[2])))
    except Exception as ex:
        log("WARNING: Could not read vertices for bbox: %s" % ex)

    if len(pts) == 0:
        try:
            for e in part_obj.edges:
                p = e.pointOn[0]
                pts.append((float(p[0]), float(p[1]), float(p[2])))
        except Exception as ex:
            die("Failed to compute bbox from geometry points (no vertices/edges usable): %s" % ex)

    xs = [p[0] for p in pts]
    ys = [p[1] for p in pts]
    zs = [p[2] for p in pts]
    low  = (min(xs), min(ys), min(zs))
    high = (max(xs), max(ys), max(zs))
    return low, high

def detect_unit_scale_from_geometry(part_obj, load_point):
    low, high = bbox_from_part_geometry(part_obj)
    y_min, y_max = float(low[1]), float(high[1])
    tip_y = y_max
    root_y = y_min

    lp_y = float(load_point[1])
    if not AUTO_UNIT_SCALE:
        return 1.0, root_y, tip_y, low, high

    if abs(lp_y) < 1e-12:
        return 1.0, root_y, tip_y, low, high

    ratio = abs(tip_y / lp_y)

    if 200.0 < ratio < 5000.0:
        scale = ratio
        log("AUTO_UNIT_SCALE: Detected geometry/load mismatch.")
        log("  tip_y ≈ %.6g, LOAD_POINT.y = %.6g, ratio ≈ %.6g -> using UNIT_SCALE=%.6g" % (tip_y, lp_y, ratio, scale))
        return scale, root_y, tip_y, low, high

    return 1.0, root_y, tip_y, low, high

def ensure_material_and_section(model, part_obj, shell_t):
    log("Ensuring material and shell section...")
    if MATERIAL_NAME not in model.materials:
        mat = model.Material(name=MATERIAL_NAME)
        mat.Elastic(table=((E, NU),))
        log("Created material: %s" % MATERIAL_NAME)

    sec_name = "BladeShellSection"
    if sec_name not in model.sections:
        model.HomogeneousShellSection(
            name=sec_name,
            material=MATERIAL_NAME,
            thicknessType=UNIFORM,
            thickness=shell_t,
            poissonDefinition=DEFAULT,
            thicknessModulus=None,
            temperature=GRADIENT,
            useDensity=OFF,
            integrationRule=SIMPSON,
            numIntPts=5,
        )
        log("Created section: %s (thickness=%.6g)" % (sec_name, shell_t))

    faces = part_obj.faces
    if len(faces) == 0:
        die("Part has no faces. Are you sure the imported blade is a shell/surface part?")

    region = regionToolset.Region(faces=faces)
    part_obj.SectionAssignment(
        region=region,
        sectionName=sec_name,
        offset=0.0,
        offsetType=MIDDLE_SURFACE,
        offsetField="",
        thicknessAssignment=FROM_SECTION,
    )
    log("Assigned section to all faces.")

def ensure_step(model):
    log("Ensuring analysis step...")
    if STEP_NAME in model.steps:
        log("Step exists: %s" % STEP_NAME)
        return
    model.StaticStep(name=STEP_NAME, previous="Initial")
    log("Created step: %s" % STEP_NAME)

def edge_avg_point(edge_obj):
    pts = edge_obj.pointOn
    xs = [float(p[0]) for p in pts]
    ys = [float(p[1]) for p in pts]
    zs = [float(p[2]) for p in pts]
    return (sum(xs)/len(xs), sum(ys)/len(ys), sum(zs)/len(zs))

def dist3(a, b):
    return math.sqrt((a[0]-b[0])**2 + (a[1]-b[1])**2 + (a[2]-b[2])**2)

def find_tip_top_edge_near_loadpoint(inst, tip_y, tol_y, load_point):
    log("Finding tip top edge nearest LOAD_POINT on tip plane...")

    tip_edges = inst.edges.getByBoundingBox(
        xMin=-1e20, xMax=1e20,
        yMin=tip_y - tol_y, yMax=tip_y + tol_y,
        zMin=-1e20, zMax=1e20
    )
    log("Tip-plane edges found: %d" % len(tip_edges))
    if len(tip_edges) == 0:
        die("No edges found near tip plane y=%.6g (tol=%.6g). Check span direction and geometry." % (tip_y, tol_y))

    top_edges = []
    for e in tip_edges:
        p = edge_avg_point(e)
        if p[2] >= 0.0:
            top_edges.append(e)
    log("Top tip edges (z>=0) candidates: %d" % len(top_edges))
    if len(top_edges) == 0:
        die("Found tip-plane edges, but none qualify as TOP (z>=0). Check coordinate system / sign convention.")

    best = None
    best_d = None
    best_p = None
    for e in top_edges:
        p = edge_avg_point(e)
        d = dist3(p, load_point)
        if (best is None) or (d < best_d):
            best = e
            best_d = d
            best_p = p

    if best is None:
        die("Failed to pick a best tip top edge (unexpected).")

    log("Selected tip-top edge: index=%s, avgPt=(%.6g, %.6g, %.6g), dist_to_LOAD_POINT≈%.6g"
        % (str(best.index), best_p[0], best_p[1], best_p[2], best_d))
    return best

def partition_edge_by_u(part_obj, inst_edge, u_param):
    log("Partitioning selected tip edge at u=%.6f ..." % float(u_param))
    try:
        part_edge = part_obj.edges[inst_edge.index]
    except Exception as ex:
        die("Could not map instance edge to part edge using index=%s: %s" % (str(inst_edge.index), ex))

    try:
        part_obj.PartitionEdgeByParam(edges=(part_edge,), parameter=float(u_param))
    except Exception as ex:
        die("PartitionEdgeByParam failed. (u=%.6f) Error: %s" % (float(u_param), ex))

    log("PartitionEdgeByParam succeeded at u=%.6f" % float(u_param))

def apply_root_pinned_bc(model, inst, root_y, tol_y):
    log("Applying pinned BC at root (y≈%.6g, tol=%.6g)..." % (root_y, tol_y))

    edges = inst.edges.getByBoundingBox(
        xMin=-1e20, xMax=1e20,
        yMin=root_y - tol_y, yMax=root_y + tol_y,
        zMin=-1e20, zMax=1e20
    )
    log("Root-plane edges found: %d" % len(edges))
    if len(edges) == 0:
        die("No root edges found near y=%.6g (tol=%.6g). Check span direction." % (root_y, tol_y))

    a = model.rootAssembly
    set_name = "SET_ROOT_EDGES"
    if set_name in a.sets:
        del a.sets[set_name]
    a.Set(name=set_name, edges=edges)

    region = a.sets[set_name]
    bc_name = "BC_ROOT_PINNED"
    if bc_name in model.boundaryConditions:
        del model.boundaryConditions[bc_name]

    model.DisplacementBC(
        name=bc_name,
        createStepName="Initial",
        region=region,
        u1=0.0, u2=0.0, u3=0.0,
        ur1=UNSET, ur2=UNSET, ur3=UNSET,
        amplitude=UNSET,
        distributionType=UNIFORM,
        fieldName="",
        localCsys=None
    )
    log("Pinned BC applied on root edges.")

def mesh_part(part_obj, mesh_seed):
    log("Meshing part (seed=%.6g)..." % float(mesh_seed))
    try:
        part_obj.setMeshControls(regions=part_obj.faces, elemShape=QUAD, technique=FREE)
    except Exception:
        pass

    part_obj.seedPart(size=float(mesh_seed), deviationFactor=0.1, minSizeFactor=0.1)

    elemType1 = mesh.ElemType(elemCode=S4R, elemLibrary=STANDARD)
    elemType2 = mesh.ElemType(elemCode=S3,  elemLibrary=STANDARD)
    part_obj.setElementType(regions=(part_obj.faces,), elemTypes=(elemType1, elemType2))

    part_obj.generateMesh()
    log("Mesh generated.")

def closest_node(nodes_obj, xyz):
    res = nodes_obj.getClosest(coordinates=xyz)
    if hasattr(res, "label"):
        return res
    if isinstance(res, (list, tuple)):
        if len(res) == 0:
            return None
        first = res[0]
        if isinstance(first, (list, tuple)) and len(first) >= 1 and hasattr(first[0], "label"):
            return first[0]
        if hasattr(first, "label"):
            return first
    return None

def apply_concentrated_force_at_loadpoint(model, inst, load_point, force_vec):
    log("Applying concentrated force at closest node to LOAD_POINT...")

    nodes = inst.nodes
    if len(nodes) == 0:
        die("No mesh nodes on instance. Meshing failed or not generated.")

    node_obj = closest_node(nodes, load_point)
    if node_obj is None:
        die("Could not find closest node to LOAD_POINT=%s. Check mesh and coordinates." % str(load_point))

    log("Closest node label=%d (target LOAD_POINT=%s)" % (int(node_obj.label), str(load_point)))

    a = model.rootAssembly
    set_name = "SET_LOAD_NODE"
    if set_name in a.sets:
        del a.sets[set_name]
    a.Set(name=set_name, nodes=inst.nodes.sequenceFromLabels((node_obj.label,)))

    load_name = "LOAD_F"
    if load_name in model.loads:
        del model.loads[load_name]

    model.ConcentratedForce(
        name=load_name,
        createStepName=STEP_NAME,
        region=a.sets[set_name],
        cf1=float(force_vec[0]),
        cf2=float(force_vec[1]),
        cf3=float(force_vec[2]),
        distributionType=UNIFORM,
        field="",
        localCsys=None
    )
    log("Concentrated force applied: (%.6g, %.6g, %.6g) N" % (force_vec[0], force_vec[1], force_vec[2]))

def make_job_and_run():
    log("Creating and running job: %s" % JOB_NAME)
    if JOB_NAME in mdb.jobs:
        del mdb.jobs[JOB_NAME]

    mdb.Job(
        name=JOB_NAME,
        model=MODEL_NAME,
        type=ANALYSIS,
        numCpus=1,
        numDomains=1,
        memory=90,
        memoryUnits=PERCENTAGE,
        explicitPrecision=SINGLE,
        nodalOutputPrecision=SINGLE,
        echoPrint=OFF,
        modelPrint=OFF,
        contactPrint=OFF,
        historyPrint=OFF
    )
    mdb.jobs[JOB_NAME].submit(consistencyChecking=OFF)
    mdb.jobs[JOB_NAME].waitForCompletion()
    log("Job completed.")

def main():
    log("===== START POST-IMPORT BLADE SCRIPT =====")
    log("USING LOAD_U=%.6f" % float(LOAD_U))
    log("USING LOAD_POINT=%s" % str(LOAD_POINT))

    model = get_model()
    part_obj = get_part(model)
    inst = get_or_create_instance(model, part_obj)

    inferred_scale, root_y_raw, tip_y_raw, bb_low, bb_high = detect_unit_scale_from_geometry(part_obj, LOAD_POINT)

    global UNIT_SCALE
    UNIT_SCALE = inferred_scale if AUTO_UNIT_SCALE else UNIT_SCALE

    shell_t_scaled = float(SHELL_T) * float(UNIT_SCALE)
    mesh_seed_scaled = float(MESH_SEED)
    load_point_scaled = (float(LOAD_POINT[0]) * float(UNIT_SCALE),
                         float(LOAD_POINT[1]) * float(UNIT_SCALE),
                         float(LOAD_POINT[2]) * float(UNIT_SCALE))

    span = float(tip_y_raw - root_y_raw)
    tol_y = max(1e-6, 1e-4 * abs(span) if abs(span) > 0 else 1e-6)

    ensure_material_and_section(model, part_obj, shell_t_scaled)
    ensure_step(model)

    curve2_tip_edge = find_tip_top_edge_near_loadpoint(inst, tip_y_raw, tol_y, load_point_scaled)

    # --- THIS IS THE CRITICAL BIT: partition at the embedded LOAD_U (from Streamlit slider)
    partition_edge_by_u(part_obj, curve2_tip_edge, LOAD_U)

    model.rootAssembly.regenerate()

    mesh_part(part_obj, mesh_seed_scaled)
    model.rootAssembly.regenerate()

    apply_root_pinned_bc(model, inst, root_y_raw, tol_y)

    # --- Apply load at closest node to the embedded LOAD_POINT (also from same u)
    apply_concentrated_force_at_loadpoint(model, inst, load_point_scaled, FORCE_VEC)

    make_job_and_run()
    log("===== END POST-IMPORT BLADE SCRIPT =====")

main()
"""

# =========================================================
# Downloads
# =========================================================

def build_design_table_excel_from_template(values_by_header: dict) -> bytes:
    if not os.path.exists(DESIGN_TABLE_TEMPLATE_PATH):
        df = pd.DataFrame([values_by_header])
        buf = BytesIO()
        with pd.ExcelWriter(buf, engine="xlsxwriter") as w:
            df.to_excel(w, sheet_name=DESIGN_TABLE_SHEET_NAME, index=False)
        return buf.getvalue()

    wb = openpyxl.load_workbook(DESIGN_TABLE_TEMPLATE_PATH)
    ws = wb[DESIGN_TABLE_SHEET_NAME] if DESIGN_TABLE_SHEET_NAME in wb.sheetnames else wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    out_row = 2

    header_to_col = {str(h): idx + 1 for idx, h in enumerate(headers) if h is not None}
    for h, v in values_by_header.items():
        if h in header_to_col:
            ws.cell(row=out_row, column=header_to_col[h]).value = v

    out = BytesIO()
    wb.save(out)
    return out.getvalue()

def build_blade_control_points_excel(
    root_c1: np.ndarray,
    root_c2: np.ndarray,
    tip_c1_3d: np.ndarray,
    tip_c2_3d: np.ndarray,
) -> bytes:
    rows_root = []
    for curve_name, cp in [("Root Curve 1", root_c1), ("Root Curve 2", root_c2)]:
        for i in range(cp.shape[0]):
            rows_root.append(
                {"Curve": curve_name, "PointIndex": i, "X": float(cp[i, 0]), "Y": 0.0, "Z": float(cp[i, 1])}
            )
    df_root = pd.DataFrame(rows_root)

    rows_tip = []
    for curve_name, cp in [("Tip Curve 1", tip_c1_3d), ("Tip Curve 2", tip_c2_3d)]:
        for i in range(cp.shape[0]):
            rows_tip.append(
                {"Curve": curve_name, "PointIndex": i, "X": float(cp[i, 0]), "Y": float(cp[i, 1]), "Z": float(cp[i, 2])}
            )
    df_tip = pd.DataFrame(rows_tip)

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        df_root.to_excel(writer, sheet_name="Root_Control_Points", index=False)
        df_tip.to_excel(writer, sheet_name="Tip_Control_Points", index=False)
    return buf.getvalue()

def build_config_template_excel() -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        default_variant_table().to_excel(writer, sheet_name="Table1_BladeVariants", index=False)
        pd.DataFrame({
            "Name": [
                "NumBlades", "ScaleX", "ScaleZ", "TwistTotalDeg", "TwistSign",
                "AbaqusLoadMag_kN", "AbaqusU",
                "AbaqusMaterialName", "AbaqusE_MPa", "AbaqusNu",
                "AbaqusMeshSeed_ModelUnits",
                "AbaqusPartName", "AbaqusJobName", "AbaqusScriptFilename",
            ],
            "Value": [
                3, 0.8, 0.8, 15.0, +1,
                1.0, 0.5,
                "BladeMaterial", 70000.0, 0.33,
                250.0,
                "Blade", "BladeJob", "abaqus.py",
            ]
        }).to_excel(writer, sheet_name="Parameters", index=False)
        pd.DataFrame({
            "Section": ["Root"]*4, "Curve": [1]*4, "PointIndex": [0, 1, 2, 3],
            "X": [1.0, 1.0, 0.8, 0.6], "Z": [0.0, 0.06, 0.08, 0.075]
        }).to_excel(writer, sheet_name="BezierControlPoints", index=False)
        pd.DataFrame({
            "PointIndex": [0, 1, 2, 3],
            "X": [-0.4, 0.3, 13.0, 2.0], "Z": [1.0, 1.0, 0.8, 0.0]
        }).to_excel(writer, sheet_name="HubControlPoints", index=False)
    return buf.getvalue()

# =========================================================
# Streamlit app
# =========================================================

def on_config_upload_change():
    uploaded_file = st.session_state.get("excel_upload", None)
    uploaded_bytes = uploaded_file.read() if uploaded_file is not None else None

    applied, warnings, variant_table = apply_config_once_if_new(uploaded_bytes)
    st.session_state.config_applied = bool(applied)
    st.session_state.config_warnings = list(warnings)
    st.session_state.variant_table = variant_table if variant_table is not None else default_variant_table()

def main():
    st.set_page_config(page_title="CAE4 Blade & Hub Tool", layout="wide")
    keep_scroll_position()
    ensure_state_defaults()

    # ----------------------------
    # Sidebar (inputs)
    # ----------------------------
    with st.sidebar:
        st.markdown("### Blade inputs")

        with st.expander("Blade geometry inputs", expanded=False):
            st.markdown("**Root – Control Points (X–Z)**")
            for i in range(4):
                col1, col2 = st.columns(2)
                with col1:
                    st.number_input(f"Curve 1 P{i} x", key=f"c1_p{i}_x", format="%.6f")
                with col2:
                    st.number_input(f"Curve 1 P{i} z", key=f"c1_p{i}_z", format="%.6f")

            st.markdown("---")
            st.markdown("**Blade variant + tip transform**")

            blade_opts = st.session_state.get("blade_options", [3, 4, 5])
            if not blade_opts:
                blade_opts = [3, 4, 5]
                st.session_state.blade_options = blade_opts
            if int(st.session_state.num_blades) not in blade_opts:
                st.session_state.num_blades = int(blade_opts[0])

            st.selectbox("Number of blades", options=blade_opts, key="num_blades")

            colA, colB = st.columns(2)
            with colA:
                st.number_input("X Scale factor", min_value=0.1, max_value=3.0, step=0.05, key="scale_x")
            with colB:
                st.number_input("Z Scale factor", min_value=0.1, max_value=3.0, step=0.05, key="scale_z")

            st.number_input(
                "Total twist from root to tip (degrees)",
                min_value=-90.0,
                max_value=90.0,
                step=0.5,
                format="%.2f",
                key="twist_total_deg",
            )
            st.session_state.twist_sign = +1.0 if float(st.session_state.twist_total_deg) >= 0 else -1.0

        with st.expander("Hub geometry inputs", expanded=False):
            st.markdown("**Hub profile – Control Points (X–Z)**")
            for i in range(4):
                c1, c2 = st.columns(2)
                with c1:
                    st.number_input(f"Hub P{i} X", key=f"hub_p{i}_x", format="%.6f")
                with c2:
                    st.number_input(f"Hub P{i} Z", key=f"hub_p{i}_z", format="%.6f")

        with st.expander("Abaqus", expanded=False):
            st.markdown("**Applied Curve 2 Tip Load**")
            st.number_input("Load (kN)", min_value=0.0, step=0.1, format="%.3f", key="abaqus_load_mag_kN")
            st.slider("Load location U (0–1)", min_value=0.01, max_value=0.99, step=0.01, key="abaqus_u")

            st.markdown("---")
            st.markdown("**Material**")
            st.text_input("Material name", key="abaqus_material_name")

            c1, c2 = st.columns(2)
            with c1:
                st.number_input("Young's modulus E (MPa)", min_value=0.0, step=1000.0, format="%.3f", key="abaqus_E_MPa")
            with c2:
                st.number_input("Poisson's ratio ν", min_value=0.0, max_value=0.499, step=0.01, format="%.4f", key="abaqus_nu")

            st.markdown("---")
            st.markdown("**Part / mesh / run naming**")
            st.text_input("Imported part name", key="abaqus_part_name")
            st.number_input("Global mesh seed size", min_value=0.0, step=1.0, format="%.6g", key="abaqus_mesh_seed_model_units")
            st.text_input("Job name", key="abaqus_job_name")
            st.text_input("Abaqus script filename", key="abaqus_script_filename")

        with st.expander("Config", expanded=False):
            st.file_uploader(
                "Upload Excel configuration (optional)",
                type=["xlsx"],
                key="excel_upload",
                on_change=on_config_upload_change,
            )

    # Config results from state
    config_applied = bool(st.session_state.get("config_applied", False))
    config_warnings = list(st.session_state.get("config_warnings", []))
    variant_table = st.session_state.get("variant_table", default_variant_table())

    # =========================================================
    # Compute geometry from session_state
    # =========================================================
    num_blades = int(st.session_state.num_blades)
    blade_length_m, blade_thickness_m = blade_length_and_thickness_from_table(num_blades, variant_table)

    scale_x = float(st.session_state.scale_x)
    scale_z = float(st.session_state.scale_z)
    twist_total_deg = float(st.session_state.twist_total_deg)
    twist_sign = +1.0 if twist_total_deg >= 0 else -1.0
    twist_deg_per_m = (abs(twist_total_deg) / blade_length_m) if blade_length_m > 0 else 0.0

    root_c1_cubic = np.array(
        [
            [st.session_state.c1_p0_x, st.session_state.c1_p0_z],
            [st.session_state.c1_p1_x, st.session_state.c1_p1_z],
            [st.session_state.c1_p2_x, st.session_state.c1_p2_z],
            [st.session_state.c1_p3_x, st.session_state.c1_p3_z],
        ],
        dtype=float,
    )
    root_c2_cubic = compute_curve2_cubic_c2(root_c1_cubic, trailing_edge=(0.0, 0.0))

    root_all_cps_top = np.vstack([root_c1_cubic, root_c2_cubic])
    root_all_cps_bottom = root_all_cps_top.copy()
    root_all_cps_bottom[:, 1] *= -1.0

    tip_all_top_2d, tip_all_top_3d = transform_tip_from_root(
        root_all_cps_top,
        scale_x=scale_x,
        scale_z=scale_z,
        twist_deg_per_m=twist_deg_per_m,
        blade_length=blade_length_m,
        twist_sign=twist_sign,
    )
    tip_all_bottom_2d, _ = transform_tip_from_root(
        root_all_cps_bottom,
        scale_x=scale_x,
        scale_z=scale_z,
        twist_deg_per_m=twist_deg_per_m,
        blade_length=blade_length_m,
        twist_sign=twist_sign,
    )

    n1 = 4
    tip_c1_top_2d = tip_all_top_2d[:n1, :]
    tip_c2_top_2d = tip_all_top_2d[n1:, :]
    tip_c1_bottom_2d = tip_all_bottom_2d[:n1, :]
    tip_c2_bottom_2d = tip_all_bottom_2d[n1:, :]

    tip_c1_top_3d = tip_all_top_3d[:n1, :]
    tip_c2_top_3d = tip_all_top_3d[n1:, :]

    x_range, z_range = compute_axis_ranges(
        root_c1_cubic, root_c2_cubic,
        tip_c1_top_2d, tip_c2_top_2d,
        tip_c1_bottom_2d, tip_c2_bottom_2d,
    )

    hub_cp_xz = np.array(
        [
            [st.session_state.hub_p0_x, st.session_state.hub_p0_z],
            [st.session_state.hub_p1_x, st.session_state.hub_p1_z],
            [st.session_state.hub_p2_x, st.session_state.hub_p2_z],
            [st.session_state.hub_p3_x, st.session_state.hub_p3_z],
        ],
        dtype=float,
    )

    x_all_root = np.concatenate([root_c1_cubic[:, 0], root_c2_cubic[:, 0]])
    chord_length = float(np.max(x_all_root) - 0.0) if x_all_root.size else 1.0

    # =========================================================
    # Abaqus load intent (THIS is what drives the downloaded script)
    # =========================================================
    load_u = float(st.session_state.abaqus_u)           # <-- user slider value
    load_mag_kN = float(st.session_state.abaqus_load_mag_kN)
    load_mag_N = load_mag_kN * 1000.0

    pt2d, tan2d = bezier_point_and_tangent(tip_c2_top_2d, load_u)
    x_u, z_u = float(pt2d[0]), float(pt2d[1])
    dxdu, dzdu = float(tan2d[0]), float(tan2d[1])

    tangent_3d = np.array([dxdu, 0.0, dzdu], dtype=float)
    normal_3d = np.array([-tangent_3d[2], 0.0, tangent_3d[0]], dtype=float)
    nmag = float(np.linalg.norm(normal_3d))
    normal_unit_3d = (normal_3d / nmag) if nmag > 0 else np.array([0.0, 0.0, -1.0])

    if float(normal_unit_3d[2]) > 0.0:
        normal_unit_3d = -normal_unit_3d

    fxN = load_mag_N * float(normal_unit_3d[0])
    fyN = load_mag_N * float(normal_unit_3d[1])
    fzN = load_mag_N * float(normal_unit_3d[2])

    dir_xz = np.array([float(normal_unit_3d[0]), float(normal_unit_3d[2])], dtype=float)
    dmag = float(np.linalg.norm(dir_xz))
    dir_xz_unit = (dir_xz / dmag) if dmag > 0 else np.array([0.0, -1.0])

    arrow_len = 0.12 * chord_length if chord_length > 0 else 0.12

    load_point_m = (x_u, float(blade_length_m), z_u)    # <-- same u
    force_vec_N = (float(fxN), float(fyN), float(fzN))

    abaqus_script_text = generate_abaqus_post_import_script_refined(
        part_name=str(st.session_state.abaqus_part_name),
        job_name=str(st.session_state.abaqus_job_name),
        script_model_name="Model-1",
        instance_name="Blade-1",
        material_name=str(st.session_state.abaqus_material_name),
        E_MPa=float(st.session_state.abaqus_E_MPa),
        nu=float(st.session_state.abaqus_nu),
        shell_thickness_m=float(blade_thickness_m),
        mesh_seed_model_units=float(st.session_state.abaqus_mesh_seed_model_units),
        load_u=float(load_u),                 # <-- embeds current slider u
        load_point_m=load_point_m,            # <-- embeds point from same u
        force_vec_N=force_vec_N,
    )

    # =========================================================
    # Build downloads
    # =========================================================
    values_by_header = {
        "Config": str(st.session_state.abaqus_job_name),
        "Num_Blades": int(num_blades),
        "Blade_Length_m": float(blade_length_m),
        "Blade_Thickness_m": float(blade_thickness_m),
        "ScaleX": float(scale_x),
        "ScaleZ": float(scale_z),
        "Twist_Total_Deg": float(twist_total_deg),
        "Twist_Deg_per_m": float(twist_deg_per_m),
        "Twist_Sign": float(twist_sign),

        "Root_C1_P0_X": float(root_c1_cubic[0, 0]), "Root_C1_P0_Z": float(root_c1_cubic[0, 1]),
        "Root_C1_P1_X": float(root_c1_cubic[1, 0]), "Root_C1_P1_Z": float(root_c1_cubic[1, 1]),
        "Root_C1_P2_X": float(root_c1_cubic[2, 0]), "Root_C1_P2_Z": float(root_c1_cubic[2, 1]),
        "Root_C1_P3_X": float(root_c1_cubic[3, 0]), "Root_C1_P3_Z": float(root_c1_cubic[3, 1]),

        "Root_C2_P0_X": float(root_c2_cubic[0, 0]), "Root_C2_P0_Z": float(root_c2_cubic[0, 1]),
        "Root_C2_P1_X": float(root_c2_cubic[1, 0]), "Root_C2_P1_Z": float(root_c2_cubic[1, 1]),
        "Root_C2_P2_X": float(root_c2_cubic[2, 0]), "Root_C2_P2_Z": float(root_c2_cubic[2, 1]),
        "Root_C2_P3_X": float(root_c2_cubic[3, 0]), "Root_C2_P3_Z": float(root_c2_cubic[3, 1]),

        "Tip_C1_P0_X": float(tip_c1_top_3d[0, 0]), "Tip_C1_P0_Y": float(tip_c1_top_3d[0, 1]), "Tip_C1_P0_Z": float(tip_c1_top_3d[0, 2]),
        "Tip_C1_P1_X": float(tip_c1_top_3d[1, 0]), "Tip_C1_P1_Y": float(tip_c1_top_3d[1, 1]), "Tip_C1_P1_Z": float(tip_c1_top_3d[1, 2]),
        "Tip_C1_P2_X": float(tip_c1_top_3d[2, 0]), "Tip_C1_P2_Y": float(tip_c1_top_3d[2, 1]), "Tip_C1_P2_Z": float(tip_c1_top_3d[2, 2]),
        "Tip_C1_P3_X": float(tip_c1_top_3d[3, 0]), "Tip_C1_P3_Y": float(tip_c1_top_3d[3, 1]), "Tip_C1_P3_Z": float(tip_c1_top_3d[3, 2]),

        "Tip_C2_P0_X": float(tip_c2_top_3d[0, 0]), "Tip_C2_P0_Y": float(tip_c2_top_3d[0, 1]), "Tip_C2_P0_Z": float(tip_c2_top_3d[0, 2]),
        "Tip_C2_P1_X": float(tip_c2_top_3d[1, 0]), "Tip_C2_P1_Y": float(tip_c2_top_3d[1, 1]), "Tip_C2_P1_Z": float(tip_c2_top_3d[1, 2]),
        "Tip_C2_P2_X": float(tip_c2_top_3d[2, 0]), "Tip_C2_P2_Y": float(tip_c2_top_3d[2, 1]), "Tip_C2_P2_Z": float(tip_c2_top_3d[2, 2]),
        "Tip_C2_P3_X": float(tip_c2_top_3d[3, 0]), "Tip_C2_P3_Y": float(tip_c2_top_3d[3, 1]), "Tip_C2_P3_Z": float(tip_c2_top_3d[3, 2]),

        "Hub_P0_X": float(hub_cp_xz[0, 0]), "Hub_P0_Z": float(hub_cp_xz[0, 1]),
        "Hub_P1_X": float(hub_cp_xz[1, 0]), "Hub_P1_Z": float(hub_cp_xz[1, 1]),
        "Hub_P2_X": float(hub_cp_xz[2, 0]), "Hub_P2_Z": float(hub_cp_xz[2, 1]),
        "Hub_P3_X": float(hub_cp_xz[3, 0]), "Hub_P3_Z": float(hub_cp_xz[3, 1]),

        "CFD_Chord": float(chord_length),
    }

    design_table_bytes = build_design_table_excel_from_template(values_by_header)
    blade_cps_bytes = build_blade_control_points_excel(
        root_c1=root_c1_cubic,
        root_c2=root_c2_cubic,
        tip_c1_3d=tip_c1_top_3d,
        tip_c2_3d=tip_c2_top_3d,
    )
    config_template_bytes = build_config_template_excel()
    abaqus_script_bytes = abaqus_script_text.encode("utf-8")

    # --- IMPORTANT practical fix for “it always looks like 0.5”:
    # Many browsers reuse the previously downloaded filename in the Downloads folder.
    # So we include the current u in the downloaded filename to make it obvious it changed.
    u_tag = f"{load_u:.2f}".replace(".", "p")
    base_name = str(st.session_state.abaqus_script_filename).strip()
    if base_name.lower().endswith(".py"):
        base_root = base_name[:-3]
    else:
        base_root = base_name
    abaqus_download_name = f"{base_root}_u{u_tag}.py"

    # =========================================================
    # Main page layout
    # =========================================================
    st.markdown(
        """
        <style>
            .block-container { padding-top: 0.5rem; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.markdown("## CAE4 – Blade & Hub Design Tool")

    tab_overview, tab_blade, tab_hub = st.tabs(["Overview", "Blade geometry", "Hub geometry"])

    with tab_overview:
        st.header("How to use this tool")
        st.markdown(
            """
            1) **Blade geometry inputs** (sidebar) define the root Curve 1 control points in the X–Z plane.  
               (The tool automatically creates root Curve 2 profile).

            2) Choose a **blade variant** then set ScaleX / ScaleZ / Twist to generate the tip profile.

            3) In **Abaqus** (sidebar), set:
               - Load magnitude (kN) and the u-location along Curve 2 on the Tip  
               - Material values (E, ν)  
               - Part name, mesh seed, job name, and script filename

            4) Use **Downloads** to export:
               - The **CATIA design table Excel**
               - The **Abaqus .py** script
            """
        )

        if config_applied:
            st.success("Config applied (as new defaults). You can now tweak values without them reverting.")
        for w in config_warnings:
            st.info(w)

        st.subheader("Abaqus values")
        df_abaqus = pd.DataFrame(
            {
                "Quantity": [
                    "Load u on Tip Curve 2 (top)",
                    "Unit direction X",
                    "Unit direction Y",
                    "Unit direction Z",
                    "Force X (kN)",
                    "Force Y (kN)",
                    "Force Z (kN)",
                    "Force magnitude (kN)",
                    "Mesh seed",
                    "Shell thickness (mm)",
                    "Part name",
                    "Job name",
                    "Script filename (download)",
                ],
                "Value": [
                    float(load_u),
                    float(normal_unit_3d[0]),
                    float(normal_unit_3d[1]),
                    float(normal_unit_3d[2]),
                    float(fxN) / 1000.0,
                    float(fyN) / 1000.0,
                    float(fzN) / 1000.0,
                    float(np.sqrt(fxN * fxN + fyN * fyN + fzN * fzN)) / 1000.0,
                    float(st.session_state.abaqus_mesh_seed_model_units),
                    float(blade_thickness_m) * 1000.0,
                    str(st.session_state.abaqus_part_name),
                    str(st.session_state.abaqus_job_name),
                    abaqus_download_name,
                ],
            }
        )
        st.table(df_abaqus)

    with tab_blade:
        st.header("Blade geometry")
        fig = make_root_and_tip_plot_only(
            root_c1_cubic, root_c2_cubic,
            tip_c1_top_2d, tip_c2_top_2d,
            tip_c1_bottom_2d, tip_c2_bottom_2d,
            x_range=x_range, z_range=z_range,
            arrow_head_xz=(x_u, z_u),
            arrow_dir_xz_unit=(float(dir_xz_unit[0]), float(dir_xz_unit[1])),
            arrow_len=float(arrow_len),
        )
        st.plotly_chart(fig, use_container_width=True)

    with tab_hub:
        st.header("Hub geometry")
        st.plotly_chart(make_hub_plot(hub_cp_xz, hub_plane_label="X–Z"), use_container_width=True)

    with st.sidebar:
        with st.expander("Downloads", expanded=False):
            st.download_button(
                "📥 CATIA design table",
                data=design_table_bytes,
                file_name="blade_hub_design_table.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.download_button(
                "📥 Control points",
                data=blade_cps_bytes,
                file_name="blade_control_points.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            # Key includes u so Streamlit always treats it as a new download target
            st.download_button(
                "📥 Abaqus script",
                data=abaqus_script_bytes,
                file_name=abaqus_download_name,
                mime="text/x-python",
                key=f"dl_abaqus_{u_tag}",
            )
            st.download_button(
                "📥 Config template",
                data=config_template_bytes,
                file_name=CONFIG_TEMPLATE_FILENAME,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

if __name__ == "__main__":
    main()
